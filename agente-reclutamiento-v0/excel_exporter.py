# ============================================================
# excel_exporter.py — Exportador a Excel con formato profesional
# Reconstruido desde bytecode (cpython-311.pyc)
# Fuente original: G:\Mi unidad\Alex\Trabajo\24-Hoy-Datastar\IA\reclutamiento\excel_exporter.py
# Compilado: 2026-02-04 16:40:43
# ============================================================

import pandas as pd
from pathlib import Path
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import config


class ExcelExporter:
    """
    Clase para exportar y gestionar el archivo Excel de resultados.
    Incluye formato profesional y tabla dinámica.
    """

    def __init__(self, output_path: str = None):
        self.output_path = Path(output_path or config.OUTPUT_FILE)
        self.sheet_name = "CVs Procesados"
        self.pivot_sheet_name = "Reporte"

    def _style_header(self, ws):
        """Aplica estilo al encabezado de la hoja"""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _adjust_column_widths(self, ws):
        """Ajusta el ancho de las columnas automáticamente"""
        column_widths = {
            "A": 15,   # Archivo
            "B": 15,   # Extensión
            "C": 18,   # Nombre
            "D": 18,   # Apellido
            "E": 30,   # Email
            "F": 18,   # Teléfono
            "G": 45,   # Educación
            "H": 25,   # Idiomas
            "I": 50,   # Trabajo más reciente
            "J": 40,   # Crecimiento
            "K": 35,   # Contradicciones
            "L": 60,   # Herramientas
            "M": 15,   # Estado
            "N": 30,   # Observaciones
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def _style_data_cells(self, ws, start_row: int, end_row: int):
        """Aplica estilo a las celdas de datos"""
        data_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        fill_even = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for row_num in range(start_row, end_row + 1):
            fill = fill_even if row_num % 2 == 0 else fill_odd
            for cell in ws[row_num]:
                cell.alignment = data_alignment
                cell.border = thin_border
                cell.fill = fill

    def _create_table(self, ws, end_row: int):
        """Crea una tabla de Excel para filtrado"""
        table_range = f"A1:N{end_row}"

        # Verificar si ya existe una tabla con ese nombre
        table = Table(
            displayName="TablaCVs",
            ref=table_range,
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style

        # Eliminar tablas existentes con el mismo nombre
        existing_tables = [t for t in ws.tables.values() if t.name == "TablaCVs"]
        if not existing_tables:
            ws.add_table(table)

    def export(self, data: List[Dict], append: bool = True) -> bool:
        """
        Exporta los datos al archivo Excel.

        Args:
            data: Lista de diccionarios con datos de candidatos
            append: Si True, agrega a datos existentes; si False, sobrescribe

        Returns:
            True si la exportación fue exitosa
        """
        if not data:
            print("⚠️  No hay datos para exportar")
            return False

        try:
            df_new = pd.DataFrame(data, columns=config.COLUMNS)

            if append and self.output_path.exists():
                df_existing = pd.read_excel(
                    self.output_path,
                    sheet_name=self.sheet_name,
                )
                df_combined = pd.concat([df_existing, df_new])
                df_combined = df_combined.drop_duplicates(keep="last")
            else:
                df_combined = df_new

            with pd.ExcelWriter(
                self.output_path,
                engine="openpyxl",
            ) as writer:
                df_combined.to_excel(
                    writer,
                    sheet_name=self.sheet_name,
                    index=False,
                )

            self._apply_styles()

            print(f"✅ Datos exportados a: {self.output_path}")
            print(f"   Total de registros: {len(df_combined)}")
            return True

        except Exception as e:
            print(f"⚠️ Error exportando a Excel: {e}")
            return False

    def _apply_styles(self):
        """Aplica estilos profesionales al archivo Excel"""
        try:
            wb = load_workbook(self.output_path)
            ws = wb[self.sheet_name]

            self._style_header(ws)
            self._adjust_column_widths(ws)

            end_row = ws.max_row
            if end_row > 1:
                self._style_data_cells(ws, 2, end_row)

            # Congelar paneles para mantener el header visible
            ws.freeze_panes = "A2"

            # Auto filtro
            ws.auto_filter.ref = ws.dimensions

            wb.save(self.output_path)

        except Exception as e:
            print(f"⚠️  Error aplicando estilos: {e}")

    def get_current_count(self) -> int:
        """Obtiene el número actual de registros en el Excel"""
        try:
            if self.output_path.exists():
                df = pd.read_excel(
                    self.output_path,
                    sheet_name=self.sheet_name,
                )
                return len(df)
        except Exception:
            pass
        return 0
